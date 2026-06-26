# -*- coding: utf-8 -*-
"""
merge_xlsx.py — 블록별 매칭 엑셀을 '하나의 통합 엑셀'로 합친다.

work 폴더의
    srwj_matched_block_001.xlsx, srwj_matched_block_003.xlsx, ...
들을 모두 읽어
    srwj_matched_all.xlsx  (시트 '매칭 결과')
한 파일로 합친다.

합친 파일에는 어느 블록의 몇 번째 턴인지 식별할 수 있도록
맨 앞에 두 열을 추가한다.

    A: 블록(archive)  … 파일명 NNN (= ROM archive 인덱스). 1,3,5,…137
    B: 턴인덱스        … 그 블록 안에서의 0-기준 순번
                          (= patch_all.py 가 턴을 맞추는 키. 원래 엑셀의 '행순서')

뒤 열들은 원본 블록 엑셀과 동일:
    대사# / 턴# / 화자ID / 속성 / 화자(번역) / 매칭률 /
    일본어(번역TXT) / 한국어(번역TXT) / 일본어(ROM 디코딩)

사용법
------
    python merge_xlsx.py                       # work → srwj_matched_all.xlsx
    python merge_xlsx.py --xlsx-dir work --out srwj_matched_all.xlsx
"""

import argparse
import glob
import os
import re
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SRC_SHEET = '매칭 결과'
DST_SHEET = '매칭 결과'

# 합친 파일의 헤더 (앞 2열이 새로 추가됨)
NEW_HEADERS = ['블록(archive)', '턴인덱스']
# 원본 블록 엑셀의 표준 헤더 (그대로 이어붙임)
SRC_HEADERS = ['대사#', '턴#', '화자ID', '속성', '화자(번역)', '매칭률',
               '일본어(번역TXT)', '한국어(번역TXT)', '일본어(ROM 디코딩)']

# 줄바꿈 표시(셀 안에 \n 이 있는 일본어/한국어 칸) 가독성용 wrap 적용 열
WRAP_COLS = {'일본어(번역TXT)', '한국어(번역TXT)', '일본어(ROM 디코딩)'}

# 블록이 바뀌는 첫 행을 살짝 강조해 경계가 보이게
BLOCK_START_FILL = PatternFill('solid', fgColor='FFF2CC')   # 옅은 노랑
HEADER_FILL      = PatternFill('solid', fgColor='4472C4')    # 파랑


def archive_idx_from_name(path):
    """파일명 srwj_matched_block_NNN.xlsx → NNN(int). 못 찾으면 None."""
    m = re.search(r'block[_-]?(\d+)', os.path.basename(path), re.I)
    return int(m.group(1)) if m else None


def main():
    ap = argparse.ArgumentParser(description='블록별 매칭 엑셀 → 통합 엑셀')
    ap.add_argument('--xlsx-dir', default='work',
                    help='블록별 엑셀이 있는 폴더 (기본 work)')
    ap.add_argument('--out', default='srwj_matched_all.xlsx',
                    help='출력 통합 엑셀 (기본 srwj_matched_all.xlsx)')
    args = ap.parse_args()

    # 1) 블록 파일 수집 (archive 인덱스 오름차순)
    files = []
    for p in glob.glob(os.path.join(args.xlsx_dir, '*.xlsx')):
        base = os.path.basename(p)
        if 'summary' in base.lower() or 'matched_all' in base.lower():
            continue
        ai = archive_idx_from_name(p)
        if ai is None:
            continue
        files.append((ai, p))
    files.sort(key=lambda x: x[0])
    if not files:
        sys.exit(f'[오류] {args.xlsx_dir} 에서 srwj_matched_block_NNN.xlsx '
                 f'파일을 찾지 못했습니다.')

    print('=' * 60)
    print(' 블록별 매칭 엑셀 → 통합 엑셀')
    print('=' * 60)
    print(f'  입력 폴더 : {args.xlsx_dir}')
    print(f'  블록 파일 : {len(files)}개 '
          f'(archive {files[0][0]} ~ {files[-1][0]})')

    # 2) 통합 워크북 생성 + 헤더
    wb = Workbook()
    ws = wb.active
    ws.title = DST_SHEET
    headers = NEW_HEADERS + SRC_HEADERS
    ws.append(headers)
    for c, _ in enumerate(headers, 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical='center')

    # 3) 블록별로 모든 데이터 행 이어붙이기
    total_rows = 0
    block_starts = []                # 통합시트에서 각 블록이 시작하는 행번호
    for ai, path in files:
        src = load_workbook(path, data_only=True, read_only=True)
        if SRC_SHEET not in src.sheetnames:
            print(f'   ⚠ {os.path.basename(path)} 에 "{SRC_SHEET}" 시트 없음 — 건너뜀')
            src.close()
            continue
        sws = src[SRC_SHEET]

        # 원본 헤더 위치를 이름으로 매핑 (열 순서가 달라도 안전)
        hdr_row = next(sws.iter_rows(min_row=1, max_row=1, values_only=True))
        name2col = {}
        for i, h in enumerate(hdr_row):
            if h is not None and str(h).strip():
                name2col[str(h).strip()] = i           # 0-기준
        # 표준 헤더가 있으면 그 순서로, 없으면 원본 순서대로
        col_order = [name2col[h] for h in SRC_HEADERS if h in name2col]
        if not col_order:                              # 헤더가 비표준이면 통째로
            col_order = list(range(len(hdr_row)))

        block_starts.append(ws.max_row + 1)
        turn_idx = 0                                   # 블록 내 0-기준 순번
        for row in sws.iter_rows(min_row=2, values_only=True):
            # 완전히 빈 행은 건너뛰지 않는다(원래 패처가 행순서로 매칭하므로
            # 모든 데이터 행을 그대로 보존해야 인덱스가 어긋나지 않는다).
            out = [ai, turn_idx] + [row[i] if i < len(row) else None
                                    for i in col_order]
            ws.append(out)
            turn_idx += 1
            total_rows += 1
        src.close()
        print(f'   archive {ai:>3}: {turn_idx:>4}턴  → 통합행 '
              f'{block_starts[-1]}~{ws.max_row}')

    # 4) 서식 (가독성)
    #    - 헤더 고정
    ws.freeze_panes = 'A2'
    #    - 줄바꿈 칸 wrap
    wrap_idx = [i for i, h in enumerate(headers, 1) if h in WRAP_COLS]
    for ci in wrap_idx:
        for cell in ws[get_column_letter(ci)]:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    #    - 블록 시작 행 강조 (블록 열만)
    for r in block_starts:
        ws.cell(r, 1).fill = BLOCK_START_FILL
        ws.cell(r, 2).fill = BLOCK_START_FILL
    #    - 열 너비
    widths = {'블록(archive)': 12, '턴인덱스': 9, '대사#': 7, '턴#': 6,
              '화자ID': 8, '속성': 7, '화자(번역)': 12, '매칭률': 8,
              '일본어(번역TXT)': 34, '한국어(번역TXT)': 34, '일본어(ROM 디코딩)': 34}
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(h, 14)
    #    - 자동 필터(블록/턴 기준 정렬·필터 편하게)
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{ws.max_row}'

    # 5) 저장
    wb.save(args.out)
    print()
    print('  [완료]')
    print(f'   합친 블록 수 : {len(block_starts)}')
    print(f'   데이터 행 수 : {total_rows}')
    print(f'   ✓ 저장: {args.out}')
    print()
    print('  패치 실행:')
    print(f'   python patch_all.py --xlsx {args.out} --expand-dict \\')
    print(f'                       --rom "Super Robot Taisen J (Japan).gba" \\')
    print(f'                       --out srwj_korean_all_s.gba')


if __name__ == '__main__':
    main()
